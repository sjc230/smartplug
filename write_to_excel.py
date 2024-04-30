import os
import openpyxl
from openpyxl.utils import get_column_letter
import threading

units_record={}
colum_name_index={"timestamp":1,"Voltage":2, "Current":3, "Power":4, "Reactive_Power":5, "Apparent_Power":6,\
                   "Power_Factor":7,  "Volt_THD":8, "Curr_THD":9, "Volt_Fund":10, "Curr_Fund":11,"Power_Energy_Acc":12, "Reactive_Energy_Acc":13, "temperature":14}

def write_new_row_to_excel(file_name, new_row, sheet_name='Sheet'):
    pass

def write_to_excel_db (file_name, id_to_find, column_name, data_to_write, sheet_name='Sheet'): 
    row=[9999999]*len(colum_name_index)
    if file_name in units_record:
        row=units_record[file_name]
        #updated current record
        if row[colum_name_index["timestamp"]-1] == id_to_find:
            row[colum_name_index[column_name]-1]=data_to_write
            units_record[file_name]=row

        # save record and refresh record with 999999
        else:  
            #luanch a new thread to save a record to excel asynchronously
            thread = threading.Thread(target=write_new_row_to_excel, args=(file_name, row))
            thread.start()
            # write_new_row_to_excel(file_name,row) 

            row=[9999999]*len(colum_name_index)
            row[colum_name_index["timestamp"]-1]=id_to_find
            row[colum_name_index[column_name]-1]=data_to_write
            units_record[file_name]=row
    #create a new record
    else:   
        row[colum_name_index["timestamp"]-1]=id_to_find
        row[colum_name_index[column_name]-1]=data_to_write
        units_record[file_name]=row



def write_to_excel(file_name, id_to_find, column_name, data_to_write, sheet_name='Sheet'):
    # 检查文件是否存在，如果不存在则创建新文件
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        worksheet['A1'] = "timestamp"
        worksheet['B1'] = "Voltage"
        worksheet['C1'] = "Current"
        worksheet['D1'] = "Power"
        worksheet['E1'] = "Reactive_Power"
        worksheet['F1'] = "Apparent_Power"
        worksheet['G1'] = "Power_Factor"
        worksheet['H1'] = "Volt_THD"
        worksheet['I1'] = "Curr_THD"
        worksheet['J1'] = "Volt_Fund"
        worksheet['K1'] = "Curr_Fund"
        worksheet['L1'] = "Power_Energy_Acc"
        worksheet['M1'] = "Reactive_Energy_Acc"
        worksheet['N1'] = "temperature"
        workbook.save(file_name)
        # return
    # print(file_name)
    # 加载现有的工作簿
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]

    # 查找与相符的列
    column_index = next((i for i, col in enumerate(worksheet.iter_cols(min_col=1, max_col=worksheet.max_column), start=1)
                        if col[0].value == column_name), None)
    # print(column_index,column_name)
    # 检查列是否存在，如果不存在则创建新列
    if column_index is None:
        return
        # max_column = worksheet.max_column
        # worksheet.insert_cols(idx=max_column)
        # new_column_letter = get_column_letter(max_column + 1)
        # worksheet.cell(row=1, column=max_column + 1, value=column_name)
        # column_index = max_column + 1

    # 查找与ID相符的行
    # exist=False
    # # for row in worksheet.iter_rows(min_row=2, values_only=True):
    # for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
    #     if row[0] == id_to_find:
    #         cell = worksheet.cell(row=row_index, column=column_index)  # 获取要写入数据的单元格对象
    #         cell.value = data_to_write  # 设置单元格的值
    #         exist=True
    #         break  # 找到匹配的行后退出循环
    # if not exist:
    #     # 如果没有找到相符的ID，则增加新行并写入数据
    #     new_row = [0] * worksheet.max_column
    #     new_row[0] = id_to_find
    #     new_row[column_index - 1] = data_to_write
    #     worksheet.append(new_row)

    # 在最后一行查找ID
    if id_to_find == worksheet.cell(row=worksheet.max_row, column=1).value :
        cell = worksheet.cell(row=worksheet.max_row, column=column_index)  # 获取要写入数据的单元格对象
        cell.value = data_to_write  # 设置单元格的值
    else:
        # 如果没有找到相符的ID，则增加新行并写入数据
        new_row = [0] * worksheet.max_column
        new_row[0] = id_to_find
        new_row[column_index - 1] = data_to_write
        worksheet.append(new_row)

    # 保存更改
    workbook.save(file_name)


def write_new_row_to_excel(file_name, new_row, sheet_name='Sheet'):
    # 检查文件是否存在，如果不存在则创建新文件
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        worksheet['A1'] = "timestamp"
        worksheet['B1'] = "Voltage"
        worksheet['C1'] = "Current"
        worksheet['D1'] = "Power"
        worksheet['E1'] = "Reactive_Power"
        worksheet['F1'] = "Apparent_Power"
        worksheet['G1'] = "Power_Factor"
        worksheet['H1'] = "Volt_THD"
        worksheet['I1'] = "Curr_THD"
        worksheet['J1'] = "Volt_Fund"
        worksheet['K1'] = "Curr_Fund"
        worksheet['L1'] = "Power_Energy_Acc"
        worksheet['M1'] = "Reactive_Energy_Acc"
        worksheet['N1'] = "temperature"
        workbook.save(file_name)
        # return
    # print(file_name)
    # 加载现有的工作簿
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]

    worksheet.append(new_row)

    # 保存更改
    workbook.save(file_name)

# ###########使用示例 ########################
# file_name = 'example.xlsx'
# id_to_find = '12345'
# column_name = 'NewColumn'
# data_to_write = 'New Data'

# write_to_excel(file_name, id_to_find, column_name, data_to_write)

# import csv
# import os

# def write_to_csv(file_name, id_to_find, column_name, data_to_write):
#     # 检查文件是否存在，如果不存在则创建新文件
#     if not os.path.exists(file_name):
#         with open(file_name, 'w', newline='', encoding='utf-8') as csvfile:
#             writer = csv.writer(csvfile)
#             writer.writerow(["timestamp", "Voltage", "Current", "Power", "Reactive_Power", "Apparent_Power",
#                              "Power_Factor", "Volt_THD", "Curr_THD", "Volt_Fund", "Curr_Fund",
#                              "Power_Energy_Acc", "Reactive_Energy_Acc", "temperature"])
#             # return

#     # 打开现有的CSV文件
#     with open(file_name, 'r', newline='', encoding='utf-8') as csvfile:
#         reader = csv.DictReader(csvfile)
#         fieldnames = reader.fieldnames

#         # 查找与相符的列
#         if column_name not in fieldnames:
#             return
#             # fieldnames.append(column_name)

#         # 查找与ID相符的行
#         exist = False
#         rows = list(reader)
#         for row in rows:
#             if row['timestamp'] == id_to_find:
#                 row[column_name] = data_to_write
#                 exist = True
#                 break

#         if not exist:
#             # 如果没有找到相符的ID，则增加新行并写入数据
#             new_row = {'timestamp': id_to_find}
#             new_row[column_name] = data_to_write
#             rows.append(new_row)

#     # 写入更新后的内容到CSV文件
#     with open(file_name, 'w', newline='', encoding='utf-8') as csvfile:
#         writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
#         writer.writeheader()
#         writer.writerows(rows)

# # ###########使用示例 ########################
# # file_name = 'example.csv'
# # id_to_find = '12345'
# # column_name = 'NewColumn'
# # data_to_write = 'New Data'

# # write_to_csv(file_name, id_to_find, column_name, data_to_write)
